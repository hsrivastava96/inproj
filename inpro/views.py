from django.shortcuts import render
from django.http import HttpResponse, JsonResponse
import os
import openpyxl
import json
from inproj.settings import STATIC_ROOT
# Create your views here.

def getData(p, f):
	file_path = os.path.join(p, f)
	wb = openpyxl.load_workbook(file_path, data_only = True)
	sheet_names = wb.sheetnames
	print(sheet_names)
	output_dict = {}
	for s_name in sheet_names:
		curr_sheet_ob = wb[s_name]
		max_row = curr_sheet_ob.max_row
		max_col = curr_sheet_ob.max_column
		rec_dict = {}
		for i in range(2, max_row + 1):
			cell_obj = curr_sheet_ob.cell(row = i, column = 1)
			r_name = cell_obj.value
			kpi_dict = {}
			for j in range(2, max_col + 1):
				cell_obj = curr_sheet_ob.cell(row = i, column = j)
				kpi_name = curr_sheet_ob.cell(row = 1, column = j).value
				kpi_dict[kpi_name] = cell_obj.value
				print(str(cell_obj.value))
			rec_dict[r_name] = kpi_dict
		output_dict[s_name] = rec_dict
	return output_dict

def excel_Download(request):
	file_path = os.path.join(STATIC_ROOT, "kpi_rec_sheet.xlsx")
	with open(file_path, "rb") as excel:
	     data = excel.read()
	response = HttpResponse(data, content_type='application/ms-excel')
	response['Content-Disposition'] = 'attachment; filename="report.xlsx"'
	return response

def test_data(request):
	# path = r"C:\Users\uic38334\Documents\interview_front_end"
	file = "kpi_rec_sheet.xlsx"
	test_data_dct = {}
	test_data_dct = getData(STATIC_ROOT, file)
	return HttpResponse(json.dumps(test_data_dct))

def index(request):
	file = "kpi_rec_sheet.xlsx"
	test_data_dct = {}
	test_data_dct = getData(STATIC_ROOT, file)
	return HttpResponse(json.dumps(test_data_dct))
